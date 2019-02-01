from django.shortcuts import render, get_object_or_404, redirect
from .models import Blog
from django.views import generic
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from . import forms
from django.contrib.auth.models import User
import openpyxl


def index(request):
    latest_user_list = User.objects.all()
    if request.user.is_authenticated:
        current_user = request.user.userprofile
        follow_user_list = list(current_user.follow.all())
        names_to_exclude = [o.username for o in follow_user_list]
        latest_user_list1 = User.objects.exclude(username__in=names_to_exclude)
        args = {
            'latest_user_list': latest_user_list,
            'latest_user_list1': latest_user_list1,
            'follow_user_list': follow_user_list,
        }
    else:
        args = {
            'latest_user_list': latest_user_list,
        }
    return render(request, 'blog/index.html', args)


def detail_view(request, pk):
    user = User.objects.get(pk=pk)
    following = list(user.userprofile.follow.all())
    followers = list(user.userprofile.followers.all())
    args = {
        'user': user,
        'follow_list': following,
        'followers': followers,
    }
    if user == request.user:
        return render(request, 'accounts/profile.html', args)
    return render(request, 'blog/detail.html', args)


class DetailBlogView(generic.DetailView):
    model = Blog
    template_name = 'blog/blog_detail.html'


@login_required(login_url="/accounts/login")
def blog_create(request):
    if request.method == 'POST':
        form = forms.CreateBlog(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.author = request.user
            instance.save()
            return redirect('blog:index')
    else:
        form = forms.CreateBlog()
    return render(request, 'blog/blog_create.html', {'form': form})


@login_required(login_url="/accounts/login")
def comment_create(request, blog_id):
    blog = get_object_or_404(Blog, pk=blog_id)
    if request.method == 'POST':
        form = forms.CreateComment(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.author = request.user
            instance.post = blog
            instance.save()
            return redirect('blog:index')
    else:
        form = forms.CreateComment()
    return render(request, 'blog/comment_create.html', {'form': form,
                                                        'blog': blog})


@login_required(login_url="/accounts/login")
def edit_blog(request, pk):
    blog = get_object_or_404(Blog, pk=pk)
    if request.user == blog.author:
        if request.method == 'POST':
            form = forms.EditBlogView(request.POST, instance=blog)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.author = request.user
                instance.save()
                return redirect('blog:blog-detail', pk=blog.pk)
        else:
            form = forms.EditBlogView(instance=blog)
        return render(request, 'blog/edit_blog.html', {'form': form, 'blog': blog})
    else:
        error_message = "ACCESS NOT ALLOWED"
        return render(request, 'blog/index.html', {'error_message': error_message})


@login_required(login_url="/accounts/login")
def delete_blog(request, pk):
    blog = get_object_or_404(Blog, pk=pk)
    if request.user == blog.author:
        blog.delete()
        return redirect('blog:index')
    else:
        error_message = "ACCESS NOT ALLOWED"
        return render(request, 'blog/index.html', {'error_message': error_message})


@staff_member_required(login_url="/accounts/login")
def create_report(request):
    wb = openpyxl.Workbook()
    for user in User.objects.all():
        ws = wb.create_sheet(user.username)
        ws.cell(row=1, column=1, value="User Profile")
        ws.cell(row=4, column=1, value="User Name:")
        ws.cell(row=4, column=3, value=user.username)
        ws.cell(row=5, column=1, value="First Name:")
        ws.cell(row=5, column=3, value=user.first_name)
        ws.cell(row=6, column=1, value="Last Name:")
        ws.cell(row=6, column=3, value=user.last_name)
        ws.cell(row=7, column=1, value="Email:")
        ws.cell(row=7, column=3, value=user.email)
        ws.cell(row=8, column=1, value="Date Joined:")
        ws.cell(row=8, column=3, value=user.date_joined)
        ws.cell(row=9, column=1, value="Number of Blogs Written")
        ws.cell(row=9, column=3, value=user.blog_set.all().count())
    del wb["Sheet"]
    wb.save('report.xlsx')
    return redirect('blog:index')










